from datetime import datetime, date
from zoneinfo import ZoneInfo
import re
import math
import pandas as pd
import streamlit as st
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.drawing.image import Image as XLImage

st.set_page_config(page_title="Enrollment Formatter", layout="centered")

try:
    logo = Image.open("header_logo.png")
    st.image(logo, width=300)
except Exception:
    pass

st.title("HCHSP Enrollment Checklist (2025–2026)")
st.markdown("Upload your **Enrollment.xlsx** file to receive a formatted version.")
uploaded_file = st.file_uploader("Upload Enrollment.xlsx", type=["xlsx"])

def coerce_to_dt(v):
    if pd.isna(v):
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            return from_excel(v)
        except Exception:
            return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                continue
    return None

def most_recent(series):
    dates, texts = [], []
    for v in pd.unique(series.dropna()):
        dt = coerce_to_dt(v)
        if dt:
            dates.append(dt)
        else:
            s = str(v).strip()
            if s:
                texts.append(s)
    if dates:
        return max(dates)
    return texts[0] if texts else None

def normalize(s: str) -> str:
    s = s.lower()
    s = re.sub(r"[\s\-\–\—_:()]+", " ", s)
    return s.strip()

def find_cols(cols, keywords):
    out = []
    for c in cols:
        if not isinstance(c, str):
            continue
        n = normalize(c)
        if any(k in n for k in keywords):
            out.append(c)
    return out

def collapse_row_values(row, col_names):
    vals = []
    for c in col_names:
        if c in row and pd.notna(row[c]) and str(row[c]).strip() != "":
            vals.append(row[c])
    if not vals:
        return None
    dts = [coerce_to_dt(v) for v in vals]
    dts = [d for d in dts if d]
    if dts:
        return max(dts)
    return str(vals[0]).strip()

def hex_to_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

PALETTE = [hex_to_rgb(h) for h in [
    "4472C4","ED7D31","A5A5A5","FFC000","5B9BD5","70AD47","C00000","7030A0",
    "00B0F0","92D050","8FAADC","FF66CC","33CCCC","9966FF","FF9933","6A5ACD",
    "2E8B57","FF7F50","1F77B4","D62728"
]]

def draw_completion_chart(names, rates, out_path):
    if not names:
        return
    left_margin = 180
    right_margin = 80
    top_margin = 100
    bottom_margin = 180
    n = len(names)
    bar_w = 40
    gap = 28
    plot_w = max(600, n * (bar_w + gap))
    plot_h = 420
    img_w = left_margin + plot_w + right_margin
    img_h = top_margin + plot_h + bottom_margin
    img = Image.new("RGB", (img_w, img_h), (255, 255, 255))
    draw = ImageDraw.Draw(img)
    try:
        title_font = ImageFont.truetype("arial.ttf", 24)
        axis_font = ImageFont.truetype("arial.ttf", 14)
        tick_font = ImageFont.truetype("arial.ttf", 12)
        label_font = ImageFont.truetype("arial.ttf", 12)
    except Exception:
        title_font = ImageFont.load_default()
        axis_font = ImageFont.load_default()
        tick_font = ImageFont.load_default()
        label_font = ImageFont.load_default()
    title = "Completion Rate of Enrollment"
    tw, th = draw.textbbox((0,0), title, font=title_font)[2:]
    draw.text(((img_w - tw)//2, 20), title, fill=(0,0,0), font=title_font)
    x0 = left_margin
    y0 = top_margin
    x1 = left_margin + plot_w
    y1 = top_margin + plot_h
    draw.rectangle([x0, y0, x1, y1], outline=(220, 220, 220), width=1)
    for p in [0.0, 0.25, 0.50, 0.75, 1.0]:
        y = y1 - int(p * plot_h)
        draw.line([(x0, y), (x1, y)], fill=(230, 230, 230), width=1)
        label = f"{int(p*100)}%"
        lw, lh = draw.textbbox((0,0), label, font=tick_font)[2:]
        draw.text((x0 - 10 - lw, y - lh//2), label, fill=(60,60,60), font=tick_font)
    y_label = "Enrollment Percentage"
    yw, yh = draw.textbbox((0,0), y_label, font=axis_font)[2:]
    y_img = Image.new("RGBA", (yh+10, yw+10), (255,255,255,0))
    y_draw = ImageDraw.Draw(y_img)
    y_draw.text((5,5), y_label, fill=(0,0,0), font=axis_font)
    y_img = y_img.rotate(90, expand=True)
    img.paste(y_img, (30, y0 + (plot_h - y_img.size[1])//2), y_img)
    x_label = "Center/Campus"
    xw, xh = draw.textbbox((0,0), x_label, font=axis_font)[2:]
    draw.text((x0 + (plot_w - xw)//2, y1 + 60), x_label, fill=(0,0,0), font=axis_font)
    max_rate = max(rates) if rates else 1.0
    y_max = max(1.0, min(1.05, max_rate + 0.05))
    for i, (name, rate) in enumerate(zip(names, rates)):
        color = PALETTE[i % len(PALETTE)]
        cx = x0 + int((i + 0.5) * (bar_w + gap))
        bh = int((rate / y_max) * plot_h)
        x_left = cx - bar_w // 2
        x_right = cx + bar_w // 2
        y_top = y1 - bh
        y_bottom = y1
        draw.rectangle([x_left, y_top, x_right, y_bottom], fill=color, outline=color)
        pct = f"{rate:.0%}"
        pw, ph = draw.textbbox((0,0), pct, font=label_font)[2:]
        label_y = max(y0 - ph + 2, y_top - ph - 4)
        draw.text((cx - pw//2, label_y), pct, fill=(20,20,20), font=label_font)
        name_text = str(name)
        nw, nh = draw.textbbox((0,0), name_text, font=tick_font)[2:]
        name_img = Image.new("RGBA", (nw+6, nh+6), (255,255,255,0))
        name_draw = ImageDraw.Draw(name_img)
        name_draw.text((3,3), name_text, fill=(40,40,40), font=tick_font)
        name_img = name_img.rotate(25, expand=True)
        nx = cx - name_img.size[0]//2
        ny = y1 + 10
        img.paste(name_img, (nx, ny), name_img)
    img.save(out_path, format="PNG")

if uploaded_file:
    wb_src = load_workbook(uploaded_file, data_only=True)
    ws_src = wb_src.active
    header_row = None
    for row in ws_src.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if isinstance(cell.value, str) and "ST: Participant PID" in cell.value:
                header_row = cell.row
                break
        if header_row:
            break
    if not header_row:
        st.error("Couldn't find 'ST: Participant PID' in the file.")
        st.stop()
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, header=header_row - 1)
    df.columns = [c.replace("ST: ", "") if isinstance(c, str) else c for c in df.columns]
    general_cutoff = datetime(2025, 5, 11)
    field_cutoff = datetime(2025, 8, 1)
    if "Participant PID" not in df.columns:
        st.error("The file is missing 'Participant PID'.")
        st.stop()
    df = (
        df.dropna(subset=["Participant PID"])
        .groupby("Participant PID", as_index=False)
        .agg(most_recent)
    )
    all_cols = list(df.columns)
    immun_cols = find_cols(all_cols, ["immun"])
    tb_cols = find_cols(all_cols, ["tb", "tuberc", "ppd"])
    lead_cols = find_cols(all_cols, ["lead", "pb"])
    if immun_cols:
        df["Immunizations"] = df.apply(lambda r: collapse_row_values(r, immun_cols), axis=1)
        df.drop(columns=[c for c in immun_cols if c in df.columns], inplace=True)
    if tb_cols:
        df["TB Test"] = df.apply(lambda r: collapse_row_values(r, tb_cols), axis=1)
        df.drop(columns=[c for c in tb_cols if c in df.columns], inplace=True)
    if lead_cols:
        df["Lead Test"] = df.apply(lambda r: collapse_row_values(r, lead_cols), axis=1)
        df.drop(columns=[c for c in lead_cols if c in df.columns], inplace=True)
    title_text = "Enrollment Checklist 2025–2026"
    central_now = datetime.now(ZoneInfo("America/Chicago"))
    timestamp_text = central_now.strftime("Generated on %B %d, %Y at %I:%M %p %Z")
    temp_path = "Enrollment_Cleaned.xlsx"
    with pd.ExcelWriter(temp_path, engine="openpyxl") as writer:
        pd.DataFrame([[title_text]]).to_excel(writer, index=False, header=False, startrow=0)
        pd.DataFrame([[timestamp_text]]).to_excel(writer, index=False, header=False, startrow=1)
        df.to_excel(writer, index=False, startrow=3)
    wb = load_workbook(temp_path)
    ws = wb.active
    filter_row = 4
    data_start = filter_row + 1
    data_end = ws.max_row
    base_max_col = ws.max_column
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{filter_row}:{get_column_letter(base_max_col)}{data_end}"
    title_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
    ts_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=base_max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=base_max_col)
    tcell = ws.cell(row=1, column=1); tcell.value = title_text
    tcell.font = Font(size=14, bold=True)
    tcell.alignment = Alignment(horizontal="center", vertical="center")
    tcell.fill = title_fill
    scell = ws.cell(row=2, column=1); scell.value = timestamp_text
    scell.font = Font(size=10, italic=True, color="555555")
    scell.alignment = Alignment(horizontal="center", vertical="center")
    scell.fill = ts_fill
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for cell in ws[filter_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    red_font = Font(color="FF0000", bold=True)

    def hdr(c):
        v = ws.cell(row=filter_row, column=c).value
        return v.strip() if isinstance(v, str) else v

    headers = [hdr(c) for c in range(1, base_max_col + 1)]
    def find_idx_exact(name):
        for i, h in enumerate(headers, start=1):
            if h == name:
                return i
        return None
    def find_idx_sub(sub):
        for i, h in enumerate(headers, start=1):
            if isinstance(h, str) and sub in h.lower():
                return i
        return None

    name_col_idx = next((i for i, h in enumerate(headers, 1) if isinstance(h, str) and "name" in h.lower()), 2)
    immun_idx = find_idx_exact("Immunizations") or find_idx_sub("immun")
    tb_idx = find_idx_exact("TB Test") or find_idx_sub("tb")
    lead_idx = find_idx_exact("Lead Test") or find_idx_sub("lead")

    scn_en_idx = (find_idx_exact("Child's Special Care Needs English") or find_idx_sub("special care needs english"))
    scn_es_idx = (find_idx_exact("Child's Special Care Needs Spanish") or find_idx_sub("special care needs spanish"))
    scn_comb_idx = scn_en_idx or scn_es_idx
    other_scn_idx = None
    if scn_comb_idx:
        other_scn_idx = scn_es_idx if scn_comb_idx == scn_en_idx else scn_en_idx
        ws.cell(row=filter_row, column=scn_comb_idx, value="Child's Special Care Needs")
        for r in range(data_start, data_end + 1):
            val_en = ws.cell(row=r, column=scn_en_idx).value if scn_en_idx else None
            val_es = ws.cell(row=r, column=scn_es_idx).value if scn_es_idx else None
            dt_en = coerce_to_dt(val_en)
            dt_es = coerce_to_dt(val_es)
            if dt_en and dt_es:
                dt = max(dt_en, dt_es)
            else:
                dt = dt_en or dt_es
            ws.cell(row=r, column=scn_comb_idx).value = dt if dt else None
        if other_scn_idx:
            ws.delete_cols(other_scn_idx, 1)
            base_max_col -= 1
            if scn_comb_idx and other_scn_idx < scn_comb_idx:
                scn_comb_idx -= 1
            headers = [hdr(c) for c in range(1, base_max_col + 1)]

    center_idx = next((i for i, h in enumerate(headers, 1) if isinstance(h, str) and ("center" in h.lower() or "campus" in h.lower() or "school" in h.lower())), None)

    for r in range(1, ws.max_row + 1):
        for c in range(1, base_max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "filtered total" in v.lower():
                ws.cell(row=r, column=c).value = None

    for r in range(data_start, data_end + 1):
        for c in range(1, base_max_col + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            cell.border = thin_border
            if val in (None, "", "nan", "NaT"):
                cell.value = "X"
                cell.font = red_font
                continue
            dt = coerce_to_dt(val)
            if immun_idx and c == immun_idx:
                if dt:
                    cell.value = dt
                    cell.number_format = "m/d/yy"
                    if dt < field_cutoff:
                        cell.font = red_font
                continue
            if tb_idx and c == tb_idx:
                if dt:
                    cell.value = dt
                    cell.number_format = "m/d/yy"
                    if dt < field_cutoff:
                        cell.font = red_font
                continue
            if scn_comb_idx and c == scn_comb_idx:
                if dt:
                    cell.value = dt
                    cell.number_format = "m/d/yy"
                    if dt < field_cutoff:
                        cell.font = red_font
                else:
                    cell.value = "X"
                    cell.font = red_font
                continue
            if lead_idx and c == lead_idx:
                if dt:
                    cell.value = dt
                    cell.number_format = "m/d/yy"
                    if dt < field_cutoff:
                        cell.font = red_font
                continue
            if dt:
                if dt < general_cutoff:
                    cell.value = "X"
                    cell.font = red_font
                else:
                    cell.value = dt
                    cell.number_format = "m/d/yy"

    width_map = {1: 16, 2: 22}
    for c in range(1, base_max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = width_map.get(c, 14)

    helper_col = base_max_col + 1
    helper_letter = get_column_letter(helper_col)
    ws.cell(row=filter_row, column=helper_col, value="VisibleFlag").font = Font(bold=True)
    anchor = f"$A${data_start}"
    for r in range(data_start, data_end + 1):
        ws.cell(row=r, column=helper_col).value = f'=SUBTOTAL(103,OFFSET({anchor},ROW()-ROW({anchor}),0))'
    ws.column_dimensions[helper_letter].hidden = True

    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=1, value="Grand Total").font = Font(bold=True)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    top_border = Border(top=Side(style="thin"))
    vis_range = f"${helper_letter}${data_start}:${helper_letter}${data_end}"
    for c in range(1, base_max_col + 1):
        if c <= name_col_idx:
            continue
        col_letter = get_column_letter(c)
        data_range = f"${col_letter}${data_start}:${col_letter}${data_end}"
        formula = f'=SUMPRODUCT(--({vis_range}=1),--({data_range}<>""),--({data_range}<>"X"))'
        cell = ws.cell(row=total_row, column=c)
        cell.value = formula
        cell.alignment = center_align
        cell.font = Font(bold=True)
        cell.border = top_border

    H_idx = 8
    R_idx = min(18, ws.max_column - 1)
    req_cols = []
    for c in range(H_idx, R_idx + 1):
        htext = hdr(c)
        if htext and str(htext).upper() != "VISIBLEFLAG":
            req_cols.append(c)

    ws_summary = wb.create_sheet(title="Center Summary")
    ws_summary.append(["Center/Campus", "Completed Students", "Total Students", "Completion Rate"])
    for c in range(1, 5):
        ws_summary.cell(row=1, column=c).font = Font(bold=True)
        ws_summary.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center")

    def value_is_complete(v):
        if v is None:
            return False
        if isinstance(v, str):
            if v.strip() == "":
                return False
            if v.strip().lower() == "x":
                return False
        return True

    center_stats = {}
    for r in range(data_start, data_end + 1):
        cname = ws.cell(row=r, column=center_idx).value if center_idx else "Unknown"
        cname = "Unknown" if cname is None or str(cname).strip() == "" else str(cname).strip()
        center_stats.setdefault(cname, {"total": 0, "completed": 0})
        center_stats[cname]["total"] += 1
        row_complete = True
        for c in req_cols:
            v = ws.cell(row=r, column=c).value
            if not value_is_complete(v):
                row_complete = False
                break
        if row_complete:
            center_stats[cname]["completed"] += 1

    sorted_centers = sorted(
        center_stats.items(),
        key=lambda kv: (0 if kv[1]["total"] == 0 else kv[1]["completed"]/kv[1]["total"]),
        reverse=True
    )

    names, rates = [], []
    row_i = 2
    for cname, stats in sorted_centers:
        total = stats["total"]
        completed = stats["completed"]
        rate = 0 if total == 0 else completed / total
        ws_summary.cell(row=row_i, column=1, value=cname)
        ws_summary.cell(row=row_i, column=2, value=completed)
        ws_summary.cell(row=row_i, column=3, value=total)
        rc = ws_summary.cell(row=row_i, column=4, value=rate)
        rc.number_format = "0%"
        names.append(str(cname))
        rates.append(rate)
        row_i += 1

    last_row = row_i - 1
    ws_summary.auto_filter.ref = f"A1:D{last_row}"
    ws_summary.column_dimensions["A"].width = 36
    ws_summary.column_dimensions["B"].width = 22
    ws_summary.column_dimensions["C"].width = 18
    ws_summary.column_dimensions["D"].width = 20

    if len(names) > 0:
        chart_path = "center_completion_chart.png"
        draw_completion_chart(names, rates, chart_path)
        img = XLImage(chart_path)
        img.width = 1100
        img.height = 600
        ws_summary.add_image(img, "F2")

    ws_scn = wb.create_sheet(title="Child's Special Care Needs Summary")
    ws_scn.append(["Center/Campus", "Completed SCN", "Total Students", "Remaining", "Completion Rate"])
    for c in range(1, 6):
        ws_scn.cell(row=1, column=c).font = Font(bold=True)
        ws_scn.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center")

    def scn_value_is_complete(v):
        if v is None:
            return False
        if isinstance(v, str):
            if v.strip() == "" or v.strip().lower() == "x":
                return False
        return True

    scn_stats = {}
    for r in range(data_start, data_end + 1):
        cname = ws.cell(row=r, column=center_idx).value if center_idx else "Unknown"
        cname = "Unknown" if cname is None or str(cname).strip() == "" else str(cname).strip()
        scn_stats.setdefault(cname, {"total": 0, "completed": 0})
        scn_stats[cname]["total"] += 1
        scn_val = ws.cell(row=r, column=scn_comb_idx).value if scn_comb_idx else None
        if scn_value_is_complete(scn_val):
            scn_stats[cname]["completed"] += 1

    sorted_scn = sorted(
        scn_stats.items(),
        key=lambda kv: (0 if kv[1]["total"] == 0 else kv[1]["completed"]/kv[1]["total"]),
        reverse=True
    )

    scn_names, scn_rates = [], []
    row_j = 2
    for cname, stats in sorted_scn:
        total = stats["total"]
        completed = stats["completed"]
        remaining = total - completed
        rate = 0 if total == 0 else completed / total
        ws_scn.cell(row=row_j, column=1, value=cname)
        ws_scn.cell(row=row_j, column=2, value=completed)
        ws_scn.cell(row=row_j, column=3, value=total)
        ws_scn.cell(row=row_j, column=4, value=remaining)
        rc = ws_scn.cell(row=row_j, column=5, value=rate)
        rc.number_format = "0%"
        scn_names.append(str(cname))
        scn_rates.append(rate)
        row_j += 1

    last_j = row_j - 1
    ws_scn.auto_filter.ref = f"A1:E{last_j}"
    ws_scn.column_dimensions["A"].width = 36
    ws_scn.column_dimensions["B"].width = 20
    ws_scn.column_dimensions["C"].width = 18
    ws_scn.column_dimensions["D"].width = 16
    ws_scn.column_dimensions["E"].width = 20

    if len(scn_names) > 0:
        scn_chart_path = "scn_completion_chart.png"
        draw_completion_chart(scn_names, scn_rates, scn_chart_path)
        scn_img = XLImage(scn_chart_path)
        scn_img.width = 1100
        scn_img.height = 600
        ws_scn.add_image(scn_img, "G2")

    final_output = "Formatted_Enrollment_Checklist.xlsx"
    wb.save(final_output)
    with open(final_output, "rb") as f:
        st.download_button("📥 Download Formatted Excel", f, file_name=final_output)

